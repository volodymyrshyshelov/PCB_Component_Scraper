"""
------------------------------------------------------------
 PCB Component Scraper
 Version: 1.0
 Release Date: 11.02.2025
 Author: Volodymyr "Zatokin" Shyshelov
 Email: volodymyr.shyshelov@gmail.com
------------------------------------------------------------

Description:
 This script is designed for automatic parsing of component 
 information from the JLCPCB website and updating the data 
 in an Excel file.

 Features:
 ✅ Parse data by part number or URL
 ✅ Automatically update the Excel file
 ✅ Preserve table formatting
 ✅ PyQt6-based GUI for user-friendly interaction
 ✅ Process logging for better tracking
 ✅ Avoid duplicates and correctly sort bulk pricing

 Requirements:
 - Python 3.8+
 - Selenium
 - pandas
 - openpyxl
 - PyQt6
 - webdriver-manager

 Installing dependencies:
 pip install selenium pandas openpyxl PyQt6 webdriver-manager

 Usage:
 1. Run the program
 2. Select an Excel file or create a new one
 3. Start parsing and wait for the data to be updated
 4. Done! The data is updated and saved with formatting
"""
import os
import time
import pandas as pd
import sys
from PyQt6.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QLabel, QProgressBar, QTextEdit
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

class ComponentScraperApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.file_path = ""

    def initUI(self):
        self.setWindowTitle("PCB Component Scraper")
        self.setGeometry(300, 300, 600, 500)

        layout = QVBoxLayout()

        self.btn_create_table = QPushButton("Створити нову таблицю за шаблоном")
        self.btn_create_table.clicked.connect(self.create_table)
        layout.addWidget(self.btn_create_table)

        self.btn_select_file = QPushButton("Обрати існуючий Excel-файл")
        self.btn_select_file.clicked.connect(self.select_file)
        layout.addWidget(self.btn_select_file)

        self.label_file_path = QLabel("Файл не обрано")
        layout.addWidget(self.label_file_path)

        self.btn_run_scraper = QPushButton("Запустити парсинг")
        self.btn_run_scraper.clicked.connect(self.run_scraper)
        self.btn_run_scraper.setEnabled(False)
        layout.addWidget(self.btn_run_scraper)

        self.progress = QProgressBar(self)
        layout.addWidget(self.progress)

        self.log_output = QTextEdit(self)
        self.log_output.setReadOnly(True)
        layout.addWidget(self.log_output)

        self.setLayout(layout)

    def create_table(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Зберегти як", "components.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            columns = ["Part Number", "Component Link", "Site Tag", "In Stock", "Unit Price", "Bulk Prices",
                       "Manufacturer", "Description", "Min Order", "Datasheet"]
            df = pd.DataFrame(columns=columns)
            df.to_excel(file_path, index=False)

            wb = load_workbook(file_path)
            ws = wb.active

            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, len(columns) + 1):
                col_letter = get_column_letter(col)
                ws[f"{col_letter}1"].font = header_font
                ws[f"{col_letter}1"].alignment = header_alignment
                ws.column_dimensions[col_letter].width = 20  

            wb.save(file_path)
            self.log(f"✅ Таблицю створено: {file_path}")

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Оберіть Excel-файл", "", "Excel Files (*.xlsx)")
        if file_path:
            self.file_path = file_path
            self.label_file_path.setText(f"Обрано файл: {file_path}")
            self.btn_run_scraper.setEnabled(True)

    def log(self, message):
        self.log_output.append(message)
        QApplication.processEvents()

    def run_scraper(self):
        if not self.file_path:
            self.log("⚠️ Файл не обрано!")
            return

        self.log("🚀 Запуск парсингу...")

        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        df = pd.read_excel(self.file_path, dtype=str)  
        total_rows = len(df)
        processed = 0

        for index, row in df.iterrows():
            link = row["Component Link"]
            part_number = row["Part Number"]

            if pd.isna(link) or link.strip() == "":
                if pd.isna(part_number) or part_number.strip() == "":
                    self.log(f"⚠️ Пропущено рядок {index + 1}: Немає посилання та парт-номери.")
                    continue
                link = f"https://jlcpcb.com/parts/componentSearch?searchTxt={part_number.strip()}"
                df.at[index, "Component Link"] = link

            df.at[index, "Site Tag"] = "jlcpcb"

            self.log(f"🔍 Обробляємо компонент: {link}")
            driver.get(link)
            time.sleep(3)

            try:
                part_number_on_page = driver.find_element(By.CSS_SELECTOR, "span.lucene_highlight_class").text

                try:
                    stock = driver.find_element(By.XPATH, "//td[contains(@class, 'el-table_1_column_9')]/div/div").text.strip()
                except:
                    stock = "Out of stock"

                try:
                    price = driver.find_element(By.CSS_SELECTOR, "div.leading-\\[20px\\] span:nth-child(2)").text.strip()
                except:
                    price = "N/A"

                try:
                    manufacturer = driver.find_element(By.XPATH, "//td[contains(@class, 'el-table_1_column_7')]/div").text.strip()
                except:
                    manufacturer = "Unknown"

                try:
                    description = driver.find_element(By.CSS_SELECTOR, "div.el-tooltip.desc-text span").text.strip()
                except:
                    description = "No description"

                try:
                    min_order_text = driver.find_element(By.XPATH, "//div[contains(@class, 'text-999999') and contains(@class, 'mb-4')]").text
                    min_order = min_order_text.split("Min:")[1].split("\n")[0].strip()
                except:
                    min_order = "N/A"

                try:
                    datasheet = driver.find_element(By.CSS_SELECTOR, "a.text-2B8CED[href$='.pdf']").get_attribute("href")
                except:
                    datasheet = ""

                bulk_prices = {}
                try:
                    bulk_blocks = driver.find_elements(By.CSS_SELECTOR, "div[data-v-04123240]")
                    for block in bulk_blocks:
                        spans = block.find_elements(By.TAG_NAME, "span")
                        if len(spans) >= 2:
                            qty = spans[0].text.strip()
                            price_bulk = spans[1].text.strip()
                            bulk_prices[qty] = price_bulk  

                    bulk_prices = {k: v for k, v in sorted(bulk_prices.items(), key=lambda item: int(item[0].replace("+", "").replace(",", "")))}
                    bulk_prices_str = ", ".join([f"{k}: {v}" for k, v in bulk_prices.items()])
                except:
                    bulk_prices_str = "N/A"

                df.at[index, "Part Number"] = part_number_on_page
                df.at[index, "Description"] = description
                df.at[index, "Manufacturer"] = manufacturer
                df.at[index, "In Stock"] = stock
                df.at[index, "Unit Price"] = price
                df.at[index, "Min Order"] = min_order
                df.at[index, "Datasheet"] = datasheet
                df.at[index, "Bulk Prices"] = bulk_prices_str

                self.log(f"✅ Дані отримані: {part_number_on_page}")

            except Exception as e:
                self.log(f"⚠️ Помилка: {e}")

            processed += 1
            self.progress.setValue(int((processed / total_rows) * 100))

        wb = load_workbook(self.file_path)
        ws = wb.active


        for i, row in df.iterrows():
            ws[f"A{i+2}"] = row["Part Number"]
            ws[f"B{i+2}"] = row["Component Link"]
            ws[f"C{i+2}"] = row["Site Tag"]
            ws[f"D{i+2}"] = row["In Stock"]
            ws[f"E{i+2}"] = row["Unit Price"]
            ws[f"F{i+2}"] = row["Bulk Prices"]
            ws[f"G{i+2}"] = row["Manufacturer"]
            ws[f"H{i+2}"] = row["Description"]
            ws[f"I{i+2}"] = row["Min Order"]
            ws[f"J{i+2}"] = row["Datasheet"]


        wb.save(self.file_path)


        self.log(f"✅ Файл оновлено: {self.file_path}")

        driver.quit()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ComponentScraperApp()
    window.show()
    sys.exit(app.exec())
