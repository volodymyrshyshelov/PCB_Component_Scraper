"""
------------------------------------------------------------
 PCB Component Scraper
 Version: 1.0
 Release Date: 11.02.2025
 Author: Volodymyr "Zatokin" Shyshelov
 Email: volodymyr.shyshelov@gmail.com
------------------------------------------------------------

Description:
 This script is designed for automatic parsing of component 
 information from the JLCPCB website and updating the data 
 in an Excel file.

 Features:
 âœ… Parse data by part number or URL
 âœ… Automatically update the Excel file
 âœ… Preserve table formatting
 âœ… PyQt6-based GUI for user-friendly interaction
 âœ… Process logging for better tracking
 âœ… Avoid duplicates and correctly sort bulk pricing

 Requirements:
 - Python 3.8+
 - Selenium
 - pandas
 - openpyxl
 - PyQt6
 - webdriver-manager

 Installing dependencies:
 pip install selenium pandas openpyxl PyQt6 webdriver-manager

 Usage:
 1. Run the program
 2. Select an Excel file or create a new one
 3. Start parsing and wait for the data to be updated
 4. Done! The data is updated and saved with formatting
"""
import os
import time
import pandas as pd
import sys
from PyQt6.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QLabel, QProgressBar, QTextEdit
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

LANGUAGES = {
    "en": {
        "title": "PCB Component Scraper",
        "create_table": "Create New Table from Template",
        "select_file": "Select Existing Excel File",
        "file_not_selected": "File not selected",
        "start_parsing": "Start Parsing",
        "log_start": "ðŸš€ Starting parsing...",
        "log_file_updated": "âœ… File updated: ",
        "switch_lang": "Switch Language"
    },
    "ua": {
        "title": "ÐŸÐ°Ñ€ÑÐµÑ€ ÐºÐ¾Ð¼Ð¿Ð¾Ð½ÐµÐ½Ñ‚Ñ–Ð² PCB",
        "create_table": "Ð¡Ñ‚Ð²Ð¾Ñ€Ð¸Ñ‚Ð¸ Ð½Ð¾Ð²Ñƒ Ñ‚Ð°Ð±Ð»Ð¸Ñ†ÑŽ Ð·Ð° ÑˆÐ°Ð±Ð»Ð¾Ð½Ð¾Ð¼",
        "select_file": "ÐžÐ±Ñ€Ð°Ñ‚Ð¸ Ñ–ÑÐ½ÑƒÑŽÑ‡Ð¸Ð¹ Excel-Ñ„Ð°Ð¹Ð»",
        "file_not_selected": "Ð¤Ð°Ð¹Ð» Ð½Ðµ Ð¾Ð±Ñ€Ð°Ð½Ð¾",
        "start_parsing": "Ð—Ð°Ð¿ÑƒÑÑ‚Ð¸Ñ‚Ð¸ Ð¿Ð°Ñ€ÑÐ¸Ð½Ð³",
        "log_start": "ðŸš€ Ð—Ð°Ð¿ÑƒÑÐº Ð¿Ð°Ñ€ÑÐ¸Ð½Ð³Ñƒ...",
        "log_file_updated": "âœ… Ð¤Ð°Ð¹Ð» Ð¾Ð½Ð¾Ð²Ð»ÐµÐ½Ð¾: ",
        "switch_lang": "Ð—Ð¼Ñ–Ð½Ð¸Ñ‚Ð¸ Ð¼Ð¾Ð²Ñƒ"
    }
}

class ComponentScraperApp(QWidget):
    

    def switch_language(self):
    # ÐŸÐµÑ€ÐµÐºÐ»ÑŽÑ‡Ð°ÐµÐ¼ ÑÐ·Ñ‹Ðº
        self.language = "ua" if self.language == "en" else "en"

    # ÐžÐ±Ð½Ð¾Ð²Ð»ÑÐµÐ¼ Ñ‚ÐµÐºÑÑ‚Ñ‹ Ð²ÑÐµÑ… ÑÐ»ÐµÐ¼ÐµÐ½Ñ‚Ð¾Ð² Ð¸Ð½Ñ‚ÐµÑ€Ñ„ÐµÐ¹ÑÐ°
        self.setWindowTitle(LANGUAGES[self.language]["title"])
        self.btn_switch_lang.setText(LANGUAGES[self.language]["switch_lang"])
        self.btn_create_table.setText(LANGUAGES[self.language]["create_table"])
        self.btn_select_file.setText(LANGUAGES[self.language]["select_file"])
        self.btn_run_scraper.setText(LANGUAGES[self.language]["start_parsing"])

        if self.file_path:
            self.label_file_path.setText(f"{LANGUAGES[self.language]['select_file']}: {self.file_path}")
        else:
            self.label_file_path.setText(LANGUAGES[self.language]["file_not_selected"])



    def __init__(self):
        super().__init__()
        self.language = "en"  # Default language
        self.initUI()
        self.file_path = ""

    def initUI(self):
        self.setWindowTitle(LANGUAGES[self.language]["title"])
        self.setGeometry(300, 300, 600, 500)

        layout = QVBoxLayout()

        self.btn_switch_lang = QPushButton(LANGUAGES[self.language]["switch_lang"])
        self.btn_switch_lang.clicked.connect(self.switch_language)
        layout.addWidget(self.btn_switch_lang)

        self.btn_create_table = QPushButton(LANGUAGES[self.language]["create_table"])
        self.btn_create_table.clicked.connect(self.create_table)
        layout.addWidget(self.btn_create_table)

        self.btn_select_file = QPushButton(LANGUAGES[self.language]["select_file"])
        self.btn_select_file.clicked.connect(self.select_file)
        layout.addWidget(self.btn_select_file)

        self.label_file_path = QLabel(LANGUAGES[self.language]["file_not_selected"])
        layout.addWidget(self.label_file_path)

        self.btn_run_scraper = QPushButton(LANGUAGES[self.language]["start_parsing"])
        self.btn_run_scraper.clicked.connect(self.run_scraper)
        self.btn_run_scraper.setEnabled(False)
        layout.addWidget(self.btn_run_scraper)

        self.progress = QProgressBar(self)
        layout.addWidget(self.progress)

        self.log_output = QTextEdit(self)
        self.log_output.setReadOnly(True)
        layout.addWidget(self.log_output)

        self.setLayout(layout)





    def create_table(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Ð—Ð±ÐµÑ€ÐµÐ³Ñ‚Ð¸ ÑÐº", "components.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            columns = ["Part Number", "Component Link", "Site Tag", "In Stock", "Unit Price", "Bulk Prices",
                       "Manufacturer", "Description", "Min Order", "Datasheet"]
            df = pd.DataFrame(columns=columns)
            df.to_excel(file_path, index=False)

            wb = load_workbook(file_path)
            ws = wb.active

            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, len(columns) + 1):
                col_letter = get_column_letter(col)
                ws[f"{col_letter}1"].font = header_font
                ws[f"{col_letter}1"].alignment = header_alignment
                ws.column_dimensions[col_letter].width = 20  

            wb.save(file_path)
            self.log(f"âœ… Ð¢Ð°Ð±Ð»Ð¸Ñ†ÑŽ ÑÑ‚Ð²Ð¾Ñ€ÐµÐ½Ð¾: {file_path}")

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "ÐžÐ±ÐµÑ€Ñ–Ñ‚ÑŒ Excel-Ñ„Ð°Ð¹Ð»", "", "Excel Files (*.xlsx)")
        if file_path:
            self.file_path = file_path
            self.label_file_path.setText(f"ÐžÐ±Ñ€Ð°Ð½Ð¾ Ñ„Ð°Ð¹Ð»: {file_path}")
            self.btn_run_scraper.setEnabled(True)

    def log(self, message):
        self.log_output.append(message)
        QApplication.processEvents()

    def run_scraper(self):
        if not self.file_path:
            self.log(f"{LANGUAGES[self.language]['log_error']} {e}")
            return

        self.log(LANGUAGES[self.language]["log_start"])

        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)

        df = pd.read_excel(self.file_path, dtype=str)  
        total_rows = len(df)
        processed = 0

        for index, row in df.iterrows():
            link = row["Component Link"]
            part_number = row["Part Number"]

            if pd.isna(link) or link.strip() == "":
                if pd.isna(part_number) or part_number.strip() == "":
                    self.log(f"âš ï¸ ÐŸÑ€Ð¾Ð¿ÑƒÑ‰ÐµÐ½Ð¾ Ñ€ÑÐ´Ð¾Ðº {index + 1}: ÐÐµÐ¼Ð°Ñ” Ð¿Ð¾ÑÐ¸Ð»Ð°Ð½Ð½Ñ Ñ‚Ð° Ð¿Ð°Ñ€Ñ‚-Ð½Ð¾Ð¼ÐµÑ€Ð¸.")
                    continue
                link = f"https://jlcpcb.com/parts/componentSearch?searchTxt={part_number.strip()}"
                df.at[index, "Component Link"] = link

            df.at[index, "Site Tag"] = "jlcpcb"

            self.log(f"{LANGUAGES[self.language]['log_processing']} {link}")
            driver.get(link)
            time.sleep(3)

            try:
                part_number_on_page = driver.find_element(By.CSS_SELECTOR, "span.lucene_highlight_class").text

                try:
                    stock = driver.find_element(By.XPATH, "//td[contains(@class, 'el-table_1_column_9')]/div/div").text.strip()
                except:
                    stock = "Out of stock"

                try:
                    price = driver.find_element(By.CSS_SELECTOR, "div.leading-\\[20px\\] span:nth-child(2)").text.strip()
                except:
                    price = "N/A"

                try:
                    manufacturer = driver.find_element(By.XPATH, "//td[contains(@class, 'el-table_1_column_7')]/div").text.strip()
                except:
                    manufacturer = "Unknown"

                try:
                    description = driver.find_element(By.CSS_SELECTOR, "div.el-tooltip.desc-text span").text.strip()
                except:
                    description = "No description"

                try:
                    min_order_text = driver.find_element(By.XPATH, "//div[contains(@class, 'text-999999') and contains(@class, 'mb-4')]").text
                    min_order = min_order_text.split("Min:")[1].split("\n")[0].strip()
                except:
                    min_order = "N/A"

                try:
                    datasheet = driver.find_element(By.CSS_SELECTOR, "a.text-2B8CED[href$='.pdf']").get_attribute("href")
                except:
                    datasheet = ""

                bulk_prices = {}
                try:
                    bulk_blocks = driver.find_elements(By.CSS_SELECTOR, "div[data-v-04123240]")
                    for block in bulk_blocks:
                        spans = block.find_elements(By.TAG_NAME, "span")
                        if len(spans) >= 2:
                            qty = spans[0].text.strip()
                            price_bulk = spans[1].text.strip()
                            bulk_prices[qty] = price_bulk  

                    bulk_prices = {k: v for k, v in sorted(bulk_prices.items(), key=lambda item: int(item[0].replace("+", "").replace(",", "")))}
                    bulk_prices_str = ", ".join([f"{k}: {v}" for k, v in bulk_prices.items()])
                except:
                    bulk_prices_str = "N/A"

                df.at[index, "Part Number"] = part_number_on_page
                df.at[index, "Description"] = description
                df.at[index, "Manufacturer"] = manufacturer
                df.at[index, "In Stock"] = stock
                df.at[index, "Unit Price"] = price
                df.at[index, "Min Order"] = min_order
                df.at[index, "Datasheet"] = datasheet
                df.at[index, "Bulk Prices"] = bulk_prices_str

                self.log(f"âœ… Ð”Ð°Ð½Ñ– Ð¾Ñ‚Ñ€Ð¸Ð¼Ð°Ð½Ñ–: {part_number_on_page}")

            except Exception as e:
                self.log(f"{LANGUAGES[self.language]['log_error']} {e}")

            processed += 1
            self.progress.setValue(int((processed / total_rows) * 100))

        wb = load_workbook(self.file_path)
        ws = wb.active


        for i, row in df.iterrows():
            ws[f"A{i+2}"] = row["Part Number"]
            ws[f"B{i+2}"] = row["Component Link"]
            ws[f"C{i+2}"] = row["Site Tag"]
            ws[f"D{i+2}"] = row["In Stock"]
            ws[f"E{i+2}"] = row["Unit Price"]
            ws[f"F{i+2}"] = row["Bulk Prices"]
            ws[f"G{i+2}"] = row["Manufacturer"]
            ws[f"H{i+2}"] = row["Description"]
            ws[f"I{i+2}"] = row["Min Order"]
            ws[f"J{i+2}"] = row["Datasheet"]


        wb.save(self.file_path)


        self.log(f"{LANGUAGES[self.language]['log_file_updated']}{self.file_path}")

        driver.quit()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ComponentScraperApp()
    window.show()
    sys.exit(app.exec())
